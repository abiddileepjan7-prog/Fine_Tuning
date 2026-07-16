"""
=========================================================
Excel/CSV Processing → Structured JSON + Timing Report
Libraries:
1. pandas
2. openpyxl
=========================================================
"""

import json
import time
import pandas as pd
from openpyxl import load_workbook

EXCEL_FILE = "Covid Dashboard.xlsx"
CSV_FILE = "Employee_Salary_Dataset.csv"
OUTPUT_JSON = "excel_csv_structured_output.json"
TIMING_JSON = "excel_csv_timing_report.json"

# ==========================================================
# Unified output schema
# ==========================================================

result = {
    "excel": {
        "source_file": EXCEL_FILE,
        "sheet_names": [],
        "sheets": []
    },
    "csv": {
        "source_file": CSV_FILE,
        "num_rows": 0,
        "num_columns": 0,
        "columns": [],
        "records": []
    }
}

# ==========================================================
# Timing tracker
# ==========================================================

timings = {}


def time_it(label, func):
    """Runs func(), records elapsed time, returns func's result."""

    start = time.perf_counter()
    output = func()
    end = time.perf_counter()

    elapsed = round(end - start, 4)

    timings[label] = {"time_seconds": elapsed}

    print(f"\n[TIMER] {label} took {elapsed} seconds")

    return output


# ==========================================================
# 1. OPENPYXL — sheet structure, images, charts
# ==========================================================

print("=" * 70)
print("1. OPENPYXL")
print("=" * 70)

sheet_media = {}


def run_openpyxl():

    wb = load_workbook(EXCEL_FILE)

    result["excel"]["sheet_names"] = wb.sheetnames

    for sheet_name in wb.sheetnames:

        ws = wb[sheet_name]

        images_count = len(ws._images) if hasattr(ws, "_images") else 0
        charts_count = len(ws._charts) if hasattr(ws, "_charts") else 0

        sheet_media[sheet_name] = {
            "images_found": images_count,
            "charts_found": charts_count
        }

        print(f"\nSheet : {sheet_name}")
        print("Rows :", ws.max_row)
        print("Columns :", ws.max_column)
        print("Images :", images_count)
        print("Charts :", charts_count)

    return len(wb.sheetnames)


sheets_found = time_it("openpyxl", run_openpyxl)

timings["openpyxl"]["sheets_found"] = sheets_found

print("\nSheet Names :", result["excel"]["sheet_names"])

# ==========================================================
# 2. PANDAS — Excel sheet data as structured records
# ==========================================================

print("\n")
print("=" * 70)
print("2. PANDAS (Excel)")
print("=" * 70)


def run_pandas_excel():

    total_rows = 0

    excel = pd.ExcelFile(EXCEL_FILE)

    for sheet_name in excel.sheet_names:

        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)

        print(f"\nSheet : {sheet_name}")
        print("Shape :", df.shape)
        print("Columns :", df.columns.tolist())

        sheet_entry = {
            "sheet_name": sheet_name,
            "num_rows": df.shape[0],
            "num_columns": df.shape[1],
            "columns": df.columns.tolist(),
            "records": df.to_dict(orient="records"),
            "images_found": sheet_media.get(sheet_name, {}).get("images_found", 0),
            "charts_found": sheet_media.get(sheet_name, {}).get("charts_found", 0)
        }

        result["excel"]["sheets"].append(sheet_entry)

        total_rows += df.shape[0]

    return total_rows


rows_pandas_excel = time_it("pandas_excel", run_pandas_excel)

timings["pandas_excel"]["rows_extracted"] = rows_pandas_excel
timings["pandas_excel"]["sheets_processed"] = len(result["excel"]["sheets"])

print("\nAll Excel sheets converted to key-value records")

# ==========================================================
# 3. PANDAS — CSV as structured records
# ==========================================================

print("\n")
print("=" * 70)
print("3. PANDAS (CSV)")
print("=" * 70)


def run_pandas_csv():

    csv_df = pd.read_csv(CSV_FILE)

    print("\nShape :", csv_df.shape)
    print("Columns :", csv_df.columns.tolist())
    print("\nHead")
    print(csv_df.head())

    result["csv"]["num_rows"] = csv_df.shape[0]
    result["csv"]["num_columns"] = csv_df.shape[1]
    result["csv"]["columns"] = csv_df.columns.tolist()
    result["csv"]["records"] = csv_df.to_dict(orient="records")

    return csv_df.shape[0]


rows_pandas_csv = time_it("pandas_csv", run_pandas_csv)

timings["pandas_csv"]["rows_extracted"] = rows_pandas_csv

print("\nCSV converted to key-value records")

# ==========================================================
# SAVE FINAL STRUCTURED JSON
# ==========================================================

with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
    json.dump(result, f, indent=2, ensure_ascii=False, default=str)

# ==========================================================
# TIMING REPORT — print to terminal + save to file
# ==========================================================

timing_report = {
    "excel_file": EXCEL_FILE,
    "csv_file": CSV_FILE,
    "timings": timings
}

print("\n")
print("=" * 70)
print("TIMING REPORT")
print("=" * 70)

print(f"\n{'Library':<18}{'Time (s)'}")
print("-" * 30)

for lib, data in timings.items():
    print(f"{lib:<18}{data['time_seconds']}")

print("\nFull Timing Report (JSON)")
print(json.dumps(timing_report, indent=2, ensure_ascii=False))

with open(TIMING_JSON, "w", encoding="utf-8") as f:
    json.dump(timing_report, f, indent=2, ensure_ascii=False)

print(f"\nTiming report saved to {TIMING_JSON}")

print("\nDone.")